#!/usr/bin/env python3
"""
convert-easa-anp-xlsx.py

Convert the EASA Aircraft Noise and Performance (ANP) database v9 Excel files
into the Plane Sound verified aircraft-noise-profile JSON.

It reads the certified LAmax Noise-Power-Distance (NPD) curves and, for each
aircraft, keeps the LOUDEST power setting per operating mode (departure /
approach) as the worst case most likely to contaminate a take. Distances are
converted from the ten standard ANP slant distances (feet) to metres.

It does NOT invent or alter any noise values. It only reshapes the rows that
exist in the EASA workbook and attaches an ICAO type designator (what ADS-B
reports) using the mapping table below. Where a mapping is uncertain it is
flagged so it can be verified rather than silently trusted.

Usage:
    pip install openpyxl
    python convert-easa-anp-xlsx.py \
        --npd   EASA_ANP_database_NPD_Data_v9.xlsx \
        --acft  EASA_ANP_database_Aircraft_v9.xlsx \
        --merge data/aircraft-noise-profiles.json \
        --out   data/aircraft-noise-profiles.json
"""

import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

FT_TO_M = 0.3048
STD_DISTANCES_FT = [200, 400, 630, 1000, 2000, 4000, 6300, 10000, 16000, 25000]

# ANP ACFT_ID / NPD_ID  ->  ICAO type designator (what ADS-B broadcasts).
# verify=True marks a mapping that is a best guess and should be confirmed.
# Where two ANP variants share one ICAO code (e.g. both A320neo engine options
# map to A20N), only ONE carries the ICAO code; the louder variant is chosen so
# the warning errs conservatively. The other is still imported (documented,
# available) but with no ICAO code so it never auto-matches live traffic.
ICAO_MAP = {
    "A320-270N": {"icao": "A20N", "note": "A320neo (PW1100G); louder of the two A320neo curves -> carries A20N"},
    "A320-250N": {"icao": None,  "note": "A320neo (LEAP-1A); A20N assigned to the louder A320-270N curve"},
    "A321-270N": {"icao": "A21N", "note": "A321neo"},
    "A330-941":  {"icao": "A339", "note": "A330-900neo"},
    "A330-743L": {"icao": "A333", "verify": True, "note": "A330 w/ RR Trent 772B (ceo engine); mapped to A330-300 as the most common A330ceo - VERIFY"},
    "A350-1041": {"icao": "A35K", "note": "A350-1000"},
    "747400RN":  {"icao": "B744", "note": "747-400"},
    "7673ER":    {"icao": "B763", "note": "767-300ER"},
    "7773ER":    {"icao": "B77W", "note": "777-300ER"},
    "7879":      {"icao": "B789", "note": "787-9"},
    "ERJ190-300": {"icao": "E290", "note": "E190-E2"},
    "ERJ190-400": {"icao": "E295", "note": "E195-E2"},
    "FAL900EX":  {"icao": "F900", "note": "Falcon 900EX (ICAO F900 covers 900 family)"},
    "G650ER":    {"icao": "GLF6", "note": "Gulfstream G650ER"},
}


def s(v):
    return str(v).strip() if v is not None else ""


def load_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [s(h) for h in rows[0]]
    return header, rows[1:]


def col(header, name):
    return header.index(name)


def build_npd_curves(npd_path):
    header, rows = load_sheet(npd_path)
    c_id = col(header, "NPD_ID")
    c_metric = col(header, "Noise Metric")
    c_mode = col(header, "Op Mode")
    c_power = col(header, "Power Setting")
    dist_cols = [(col(header, "L_%dft" % ft), ft) for ft in STD_DISTANCES_FT]

    # For each (NPD_ID, Op Mode) keep the LAmax row with the highest power.
    loudest = {}
    for r in rows:
        if s(r[c_metric]) != "LAmax":
            continue
        key = (s(r[c_id]), s(r[c_mode]))
        power = r[c_power]
        if power is None:
            continue
        prev = loudest.get(key)
        if prev is None or power > prev[c_power]:
            loudest[key] = r

    curves = {}
    for (npd_id, mode), r in loudest.items():
        points = []
        for ci, ft in dist_cols:
            v = r[ci]
            if isinstance(v, (int, float)):
                points.append({"distM": round(ft * FT_TO_M), "dba": round(float(v), 2)})
        if points:
            curves.setdefault(npd_id, {})[mode] = points
    return curves


def load_descriptions(acft_path):
    header, rows = load_sheet(acft_path)
    c_id = col(header, "NPD_ID")
    c_desc = col(header, "Description")
    c_app = col(header, "Approach Spectral Class ID")
    c_dep = col(header, "Departure Spectral Class ID")
    out = {}
    for r in rows:
        out[s(r[c_id])] = {
            "desc": s(r[c_desc]),
            "appClass": r[c_app],
            "depClass": r[c_dep],
        }
    return out


def make_profiles(curves, descriptions):
    profiles = []
    flagged = []
    for npd_id, modes in sorted(curves.items()):
        dep = modes.get("D")
        app = modes.get("A")
        thrust = []
        # Departure curve also serves climb and cruise lookups in the engine.
        if dep:
            thrust.append({"setting": "departure", "points": dep})
            thrust.append({"setting": "climb", "points": dep})
            thrust.append({"setting": "cruise", "points": dep})
        if app:
            thrust.append({"setting": "approach", "points": app})
        if not thrust:
            continue

        mapping = ICAO_MAP.get(npd_id, {})
        icao = mapping.get("icao")
        codes = [icao] if icao else []
        if mapping.get("verify"):
            flagged.append((npd_id, icao, mapping.get("note", "")))

        info = descriptions.get(npd_id, {})
        profile = {
            "sourceType": "verified",
            "label": npd_id,
            "aircraftTypeCodes": codes,
            "anpId": npd_id,
            "confidence": 0.85,
            "npd": {"refMetric": "Lmax_dBA", "thrust": thrust},
            "engineLabel": info.get("desc", ""),
            "noiseIdentifier": npd_id,
            "mappingNote": mapping.get("note", ""),
        }
        if info.get("appClass") is not None:
            profile["spectralApproachClass"] = int(info["appClass"])
        if info.get("depClass") is not None:
            profile["spectralDepartureClass"] = int(info["depClass"])
        profiles.append(profile)
    return profiles, flagged


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--npd", required=True)
    ap.add_argument("--acft", required=True)
    ap.add_argument("--merge", default=None, help="existing profiles JSON to keep (types not in v9)")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    curves = build_npd_curves(args.npd)
    descriptions = load_descriptions(args.acft)
    v9_profiles, flagged = make_profiles(curves, descriptions)

    v9_anp_ids = {p["anpId"] for p in v9_profiles}
    v9_icao = {c for p in v9_profiles for c in p["aircraftTypeCodes"]}

    merged = list(v9_profiles)
    kept = []
    if args.merge:
        try:
            existing = json.load(open(args.merge, encoding="utf-8-sig"))
            for p in existing.get("profiles", []):
                # Keep prior verified profiles that v9 does NOT supersede
                # (different ANP id AND no ICAO-code collision).
                pid = p.get("anpId")
                pcodes = set(p.get("aircraftTypeCodes", []))
                if pid in v9_anp_ids:
                    continue
                if pcodes & v9_icao:
                    continue
                kept.append(p)
        except FileNotFoundError:
            pass
    merged.extend(kept)

    out = {
        "schema": "plane-sound-aircraft-noise-profiles-v1",
        "mode": "verified-data-required",
        "sources": [
            {
                "name": "EASA Aircraft Noise and Performance (ANP) database v9 - NPD LAmax",
                "metric": "LAmax",
                "note": "Certified noise-power-distance curves. Loudest power setting per operating mode (worst case). Converted by scripts/convert-easa-anp-xlsx.py.",
                "license": "EASA ANP - see EASA ANP terms of use",
            },
            {
                "name": "EUROCONTROL/EASA ANP (legacy) - retained types not present in v9",
                "metric": "LAmax",
                "note": "Earlier verified curves kept for aircraft the v9 set does not cover (e.g. A320ceo).",
            },
        ],
        "profiles": merged,
    }

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2)

    # Report to stderr so stdout/out file stays clean.
    print("Wrote %d verified profiles (%d from v9, %d retained legacy)."
          % (len(merged), len(v9_profiles), len(kept)), file=sys.stderr)
    mapped = [(p["anpId"], p["aircraftTypeCodes"]) for p in v9_profiles]
    print("v9 ICAO mappings:", file=sys.stderr)
    for anp, codes in mapped:
        print("   %-12s -> %s" % (anp, codes or "(no auto-match)"), file=sys.stderr)
    if flagged:
        print("\nFLAGGED for verification:", file=sys.stderr)
        for anp, icao, note in flagged:
            print("   %-12s -> %s   %s" % (anp, icao, note), file=sys.stderr)


if __name__ == "__main__":
    main()
